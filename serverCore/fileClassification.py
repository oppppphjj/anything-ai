import fitz
import openpyxl
import xlrd
from docx import Document


class Datapreprocessing:
    # 文件类型分类
    def fileClassification(self, url):
        if url.endswith('.docx'):
            text = self.extract_text_from_docx(url)
            return text
        elif url.endswith('.pdf'):
            text = self.extract_text_from_pdf(url)
            return text
        elif url.endswith('.xlsx'):
            text = self.extract_text_from_xlsx(url)
            return text
        elif url.endswith('.xls'):
            text = self.extract_text_from_xls(url)
            return text
        elif url.endswith('.txt'):
            text = self.extract_text_from_txt(url)
            return text
        else:
            text = '无法识别该文件'
            return text

    # word提取文本
    def extract_text_from_docx(self,file_path):
        doc = Document(file_path)
        text = ''
        for para in doc.paragraphs:
            text += para.text + '\n'
        return text

    # pdf提取文本
    def extract_text_from_pdf(self,file_path):
        pdf_document = fitz.open(file_path)
        text = ''
        for page_num in range(pdf_document.page_count):
            page = pdf_document[page_num]
            text += page.get_text()
        return text

    # 新excel文件提取文本
    def extract_text_from_xlsx(self,file_path):
        text = ""
        try:
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            text += str(cell.value) + " "
        except Exception as e:
            print(f"An error occurred: {e}")

        return text

    # 旧excel文件提取文本
    def extract_text_from_xls(self,file_path):
        text = ""
        try:
            workbook = xlrd.open_workbook(file_path)
            for sheet in workbook.sheets():
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        cell_value = sheet.cell(row, col).value
                        if cell_value:
                            text += str(cell_value) + " "
        except Exception as e:
            print(f"An error occurred: {e}")

        return text

    # txt文件提取纯文本
    def extract_text_from_txt(self,file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                text = file.read()
                return text
        except Exception as e:
            print(f"An error occurred: {e}")
            return None